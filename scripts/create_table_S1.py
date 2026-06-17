#!/usr/bin/env python3
"""Create table_S1.xlsx with 3 sheets from source data.

Sheet 1: Spearman correlations between 15 biological features and Procrustes rigidity
Sheet 2: Cross-atlas ranking consistency master table
Sheet 3: Three-species Procrustes results summary

Source data:
  - analysis/biological_predictors/univariate_correlations.csv
  - analysis/cross_reference/master_ranking_table.csv
  - output/phase2/scaled_35types/procrustes_results_35.json
  - output/macaque_pipeline/reconstruction_qu12_results.json
  - analysis/mouse_lemur/procrustes_results.json
"""

import csv
import json
import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

BASE = str(Path(__file__).resolve().parent.parent)
OUT = os.path.join(BASE, "docs/supplementary_materials/table_S1.xlsx")

HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
HEADER_FONT = Font(name="Arial", bold=True, size=10)
DATA_FONT = Font(name="Arial", size=10)
WRAP = Alignment(wrap_text=True, vertical="top")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def style_header(ws, row, n_cols):
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = WRAP
        cell.border = THIN_BORDER


def style_data(ws, row, n_cols):
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = DATA_FONT
        cell.alignment = WRAP
        cell.border = THIN_BORDER


def auto_width(ws, min_w=10, max_w=40):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        lengths = []
        for cell in col:
            if cell.value:
                lines = str(cell.value).split("\n")
                lengths.append(max(len(l) for l in lines))
        best = max(lengths) + 2 if lengths else min_w
        ws.column_dimensions[letter].width = min(max(best, min_w), max_w)


# ── Sheet 1: Biological predictors ──────────────────────────────────

def build_sheet1(wb):
    ws = wb.active
    ws.title = "Biological Predictors"

    headers = ["Feature", "Spearman_rho", "p_value", "n", "Notes"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header(ws, 1, len(headers))

    csv_path = os.path.join(BASE, "analysis/biological_predictors/univariate_correlations.csv")
    with open(csv_path) as f:
        reader = csv.DictReader(f)
        for i, row in enumerate(reader, 2):
            rho = float(row["spearman_rho"])
            p = float(row["p_value"])
            n = int(row["n"])

            # Determine significance note
            if p < 0.01:
                note = "Significant (p < 0.01)"
            elif p < 0.05:
                note = "Significant (p < 0.05)"
            elif p < 0.10:
                note = "Marginal (p < 0.10)"
            else:
                note = ""

            ws.cell(row=i, column=1, value=row["feature"])
            ws.cell(row=i, column=2, value=round(rho, 4))
            ws.cell(row=i, column=3, value=round(p, 4) if p >= 0.0001 else f"{p:.2e}")
            ws.cell(row=i, column=4, value=n)
            ws.cell(row=i, column=5, value=note)
            style_data(ws, i, len(headers))

    # Format p-value column
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=2):
        for cell in row:
            cell.number_format = "0.0000"

    auto_width(ws)
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["E"].width = 25


# ── Sheet 2: Cross-atlas ranking consistency ─────────────────────────

def build_sheet2(wb):
    ws = wb.create_sheet("Cross-Atlas Ranking")

    headers = [
        "Cell_type", "Primary_rank", "Sun2023_rank", "PanSci_rank",
        "CellHint_rank", "PanCensus_rank", "Macaque_rank",
        "Mouse_lemur_rank", "n_replications"
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header(ws, 1, len(headers))

    csv_path = os.path.join(BASE, "analysis/cross_reference/master_ranking_table.csv")
    with open(csv_path) as f:
        reader = csv.DictReader(f)
        rows = list(reader)

    # Filter to types with n_replications >= 2
    rows_2plus = [r for r in rows if int(r["n_replications_present"]) >= 2]
    # Also include all 35 for completeness; mark n_replications
    for i, row in enumerate(rows, 2):
        ws.cell(row=i, column=1, value=row["cell_type"])
        ws.cell(row=i, column=2, value=int(row["primary_rank"]))

        for col_idx, key in enumerate([
            "Sun2023_rank", "PanSci_rank", "CellHint_rank",
            "Pan_Census_rank", "Macaque_rank", "Mouse_lemur_rank"
        ], 3):
            val = row.get(key, "")
            if val and val.strip():
                ws.cell(row=i, column=col_idx, value=round(float(val)))
            else:
                ws.cell(row=i, column=col_idx, value="--")

        ws.cell(row=i, column=9, value=int(row["n_replications_present"]))
        style_data(ws, i, len(headers))

    auto_width(ws)
    ws.column_dimensions["A"].width = 45


# ── Sheet 3: Three-species Procrustes summary ────────────────────────

def build_sheet3(wb):
    ws = wb.create_sheet("Three-Species Summary")

    headers = ["Species_pair", "Divergence_Mya", "n_types", "n_genes",
               "obs_null_ratio", "p_value", "Notes"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header(ws, 1, len(headers))

    # Human-mouse (primary 35-type)
    with open(os.path.join(BASE, "output/phase2/scaled_35types/procrustes_results_35.json")) as f:
        hm = json.load(f)
    hm_dist = hm["procrustes"]["distance"]
    hm_null_med = hm["permutation_test"]["null_distribution_summary"]["median"]
    hm_obs_null = hm_dist / hm_null_med
    hm_n_genes = hm.get("n_genes_input", 16959)

    # Human-macaque (Qu-only 12-type canonical; was deprecated 20-type RIRA-mixed)
    with open(os.path.join(BASE, "output/macaque_pipeline/reconstruction_qu12_results.json")) as f:
        ha = json.load(f)

    # Human-mouse lemur
    with open(os.path.join(BASE, "analysis/mouse_lemur/procrustes_results.json")) as f:
        hl = json.load(f)

    species_data = [
        {
            "pair": "Human-mouse",
            "div": 90,
            "n_types": 35,
            "n_genes": hm_n_genes,
            "obs_null": round(hm_obs_null, 3),
            "p": "<1e-6",  # 1,000,000-perm headline; procrustes_results_35.json stores the 10k-perm floor
            "notes": "Primary analysis (Tabula Sapiens vs Tabula Muris Senis)"
        },
        {
            "pair": "Human-macaque",
            "div": 25,
            "n_types": len(ha["types_included"]),
            "n_genes": ha["gene_space"],
            "obs_null": round(ha["permutation_test"]["obs_null_ratio_median"], 3),
            "p": ha["permutation_test"]["p_value"],
            "notes": "Crab-eating macaque (Macaca fascicularis)"
        },
        {
            "pair": "Human-mouse lemur",
            "div": hl["divergence_mya"],
            "n_types": hl["n_types"],
            "n_genes": hl["gene_space"],
            "obs_null": round(hl["permutation_test"]["obs_null_ratio"], 3),
            "p": hl["permutation_test"]["p_value"],
            "notes": "Gray mouse lemur (Microcebus murinus)"
        },
    ]

    for i, sp in enumerate(species_data, 2):
        ws.cell(row=i, column=1, value=sp["pair"])
        ws.cell(row=i, column=2, value=sp["div"])
        ws.cell(row=i, column=3, value=sp["n_types"])
        ws.cell(row=i, column=4, value=sp["n_genes"])
        ws.cell(row=i, column=5, value=sp["obs_null"])
        p_val = sp["p"]
        ws.cell(row=i, column=6, value=p_val if isinstance(p_val, str) else (f"{p_val:.4f}" if p_val >= 0.0001 else f"{p_val:.2e}"))
        ws.cell(row=i, column=7, value=sp["notes"])
        style_data(ws, i, len(headers))

    auto_width(ws)
    ws.column_dimensions["G"].width = 50


def main():
    wb = Workbook()
    build_sheet1(wb)
    build_sheet2(wb)
    build_sheet3(wb)

    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    wb.save(OUT)
    print(f"Saved table_S1.xlsx to {OUT}")

    # Verify
    from openpyxl import load_workbook
    wb2 = load_workbook(OUT)
    for name in wb2.sheetnames:
        ws = wb2[name]
        print(f"  Sheet '{name}': {ws.max_row} rows x {ws.max_column} cols")


if __name__ == "__main__":
    main()
