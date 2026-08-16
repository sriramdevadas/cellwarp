#!/usr/bin/env python3
"""Supplementary Table synthesis-pass edits.

Canonical targets: docs/supplementary_materials/
Staging copies:    docs/submission/figures_for_review/ (kept in sync)

Edits applied:
  Table S1:
    Sheet 3 (Three-Species Summary):
      - h-m row p-value 1.00e-04 → < 10⁻⁶ (1M permutations, per results_1M.json)
      - h-macaque row: n=20→12, obs/null=0.841→0.811, p=0.0002→0.0043,
        Notes → "Qu et al. 2022 only; symmetric preprocessing (re-evaluated)"
      - Add preprocessing note row at bottom
    Sheet 1 (Biological Predictors):
      - Add footer notes: descriptive vs inferential clarification (14 rows
        descriptive, progenitor only counts as inferential T40); un-logged
        cell count disclosure per Part 1.3.

  Table S2:
    Sheet 1 (Simulation Parameters):
      - Row 17 ('null_p_value_mean') description: add diagnostic context
      - Row 12 ('detection_power ... signal=3.0') description: clarify
      - Append bootstrap metadata: n_iterations, seed, subsample_fraction
    Sheet 2 (Bootstrap Ranking CIs):
      - Round floating-point artifacts (rows 24, 27) to clean numbers

  Table S3 (CellHint rank reversal):
    - Append legend row explaining CH Tissue Count availability (5/15 with
      per-type values for LOW-confidence rows; source note for the
      Spearman ρ = -0.526 computation)
    - Append confidence threshold definitions

  Table S4 (Progressive harmonization):
    - Rename rows to Level 0/1/2/3 convention
    - Prepend pre-Level-0 row showing the original ρ = -0.386 (Table S3
      unharmonized comparison)

  Table S5 (Cell type matching):
    - Move plasma cell row (rank 12) from bottom to correct rank position
      between MSC (rank 13, row 24) and MSC of adipose (rank 11, row 25)
    - Marker completeness NOT addressed here (requires canonical marker
      list; 10 rows remain "Cell Ontology + marker gene expression")

  Table S6 (CPC1):
    - Sort Sheet 2 by rigidity rank (currently alphabetical by cell_type)
    - Add rigidity_rank column to Sheet 2 for self-containment
    - Sheet 1: no content changes (legend-level caveats belong in PDF)

  Also: docs/submission/key_resources_table.md
    - Fix Qu scRNA-seq GEO accession GSE196791 → GSE196792 (Decision C).
"""
from __future__ import annotations
import csv
from pathlib import Path
from copy import copy

import openpyxl

DOCS = Path(__file__).resolve().parent.parent / "docs"
CANON = DOCS / "supplementary_materials"
STAGE = DOCS / "submission/figures_for_review"


# ------------------------------ Table S1 ---------------------------------

def edit_table_s1(xlsx_path: Path):
    wb = openpyxl.load_workbook(xlsx_path)

    # Sheet 3: Three-Species Summary
    ws = wb["Three-Species Summary"]
    # Row 2: h-m primary — p-value and notes update
    ws.cell(2, 6).value = "< 1e-6"
    ws.cell(2, 7).value = (
        "Primary analysis (Tabula Sapiens vs Tabula Muris Senis); "
        "1,000,000 permutations; raw → normalize_total(1e4) → log1p"
    )
    # Row 3: h-macaque — K12 update
    ws.cell(3, 3).value = 12
    ws.cell(3, 4).value = 13927
    ws.cell(3, 5).value = 0.811
    ws.cell(3, 6).value = 0.0043
    ws.cell(3, 7).value = (
        "Qu et al. 2022 (M. fascicularis) only; 12 types (Qu-only source "
        "with symmetric raw-count preprocessing); 10,000 permutations; "
        "supersedes deprecated RIRA+Qu combined design (n=20, 0.841, 0.0002)"
    )
    # Row 4: lemur — verify values; also clarify
    ws.cell(4, 7).value = (
        "Tabula Microcebus (Ezran et al. 2025); 10,000 permutations; floor p"
    )

    # Append a preprocessing-convention note row at the bottom
    last = ws.max_row
    ws.cell(last + 2, 1).value = "Note"
    ws.cell(last + 2, 2).value = None
    ws.cell(last + 2, 7).value = (
        "All three analyses use symmetric preprocessing: raw UMI counts → "
        "sc.pp.normalize_total(target_sum=1e4) → sc.pp.log1p, applied "
        "identically to human and macaque/lemur cells. 2,000 cells/type "
        "subsample cap (seed=42). Gene spaces differ (three-way vs pairwise "
        "ortholog intersections) but are reported per row."
    )

    # Sheet 1: Biological Predictors — append footer notes
    ws = wb["Biological Predictors"]
    last = ws.max_row
    ws.cell(last + 2, 1).value = "Note (descriptive vs inferential)"
    ws.cell(last + 2, 2).value = None
    ws.cell(last + 2, 5).value = (
        "These 15 correlations are descriptive feature comparisons computed "
        "for completeness; only the progenitor predictor (T40 in Table 1) is "
        "featured in main analyses as an inferential test. The remaining 14 "
        "are reported here without significance testing."
    )
    # Un-logged cell count disclosure
    last2 = ws.max_row
    ws.cell(last2 + 2, 1).value = "Note (cell count)"
    ws.cell(last2 + 2, 5).value = (
        "The 'Log min cell count' row uses log-transformed values (ρ=−0.033). "
        # "and Figure 6B" removed (D27): this manuscript has five figure captions
        # (Fig 1-5) and zero occurrences of "Fig 6"/"Figure 6" in manuscript_combined.txt,
        # S1_Text.txt or S2_Text.txt. The reference pointed into the seven-figure PLOS
        # Computational Biology version, which is what Zenodo CODE v1 still carries.
        "An un-logged cell count correlation (ρ=0.052, p=0.768, n=35) is "
        "reported separately in Table 1 (T36); the two tests "
        "use different transformations of the same underlying variable."
    )

    wb.save(xlsx_path)
    print(f"  S1 edited: {xlsx_path}")


# ------------------------------ Table S2 ---------------------------------

def edit_table_s2(xlsx_path: Path):
    wb = openpyxl.load_workbook(xlsx_path)

    # Sheet 1: Simulation Parameters
    ws = wb["Simulation Parameters"]
    # Row 12: detection_power clarification
    for r in range(2, ws.max_row + 1):
        k = ws.cell(r, 1).value
        if isinstance(k, str) and k.startswith("detection_power"):
            ws.cell(r, 3).value = (
                "Detection rate under the planted signal. Signal strength = 3.0 "
                "used as the nearest simulated setpoint to the calibrated "
                "strength (3.68); the 0.68 gap has negligible effect on "
                "detection power at n=35 (>99% under either setpoint)."
            )
        elif k == "null_p_value_mean":
            ws.cell(r, 3).value = (
                "Mean p-value under null (expected: 0.5). Diagnostic check: "
                "null distribution p-value mean should approximate 0.5 for a "
                "well-calibrated permutation test."
            )
    # Append bootstrap metadata as three rows
    last = ws.max_row
    extra = [
        ("bootstrap_n_iterations", 1000, "Bootstrap iterations for per-type ranking CIs (Sheet 2)"),
        ("bootstrap_subsample_fraction", 0.5, "Fraction of cells resampled per type per iteration"),
        ("bootstrap_random_seed", 42, "Global seed for bootstrap resampling and permutation tests"),
    ]
    for i, (k, v, d) in enumerate(extra, start=1):
        ws.cell(last + i, 1).value = k
        ws.cell(last + i, 2).value = v
        ws.cell(last + i, 3).value = d

    # Sheet 2: round CI_lower and CI_width floating-point artifacts
    ws = wb["Bootstrap Ranking CIs"]
    for r in range(2, ws.max_row + 1):
        for col in (3, 5):  # CI_lower, CI_width
            v = ws.cell(r, col).value
            if isinstance(v, float) and v != round(v):
                if abs(v - round(v)) < 0.05:
                    ws.cell(r, col).value = round(v)
                else:
                    # Keep meaningful halves etc. but this data should all be int-like
                    ws.cell(r, col).value = round(v, 2)

    wb.save(xlsx_path)
    print(f"  S2 edited: {xlsx_path}")


# ------------------------------ Table S3 ---------------------------------

def edit_table_s3(csv_path: Path):
    """No longer edits S3. The legend this appended was wrong; corrected by removal.

    It appended a five-line legend duplicating the manuscript's own S3 Table caption.
    The caption already carries all three statements, correctly, so nothing is lost --
    and each of the three departed from it, two of them falsifiably against the table's
    own contents:

      * "The Rank Difference column uses the second (matched 15-type) rank" -- FALSE.
        The caption says the first column. Checked over all 15 rows:
        Rank Difference == Primary(of 15) - CellHint holds for 12 (3 ties where the two
        columns are equal); == Primary(matched 15) - CellHint holds for 0.
        e.g. hepatocyte 13-1=+12 (matches) vs 14-1=13; neutrophil 3-13=-10 vs 9-13=-4.

      * "tissue count values tabulated here for the 5 LOW-confidence rows ... Remaining
        10 MODERATE/HIGH rows ... not separately tabulated" -- FALSE. All 15 rows carry a
        CellHint Tissue Count, and the caption says so: "tissue counts for all 15 rows".

      * "MODERATE = |rank diff| 5-7" -- the caption says 5-8. No row has |diff| = 8, so
        the data does not discriminate, but 5-7 leaves 8 unclassifiable between MODERATE
        and LOW (>= 9). The submitted caption is the authority.

    It also cited "Table 1 (T58)"; the submitted texts name that workbook "S13 Table"
    (6 occurrences) and never "Table 1" (0). T58 itself is correct: S13 Table gives
    rho = -0.526, p = 0.044, n = 15, Figure/Table "S3 Table".

    The deposited table_S3.csv never received this block and is correct as it stands.
    Deliberately does not open or rewrite the file, so no future run can reintroduce it.
    """
    print(f"  S3: legend removed (see docstring); {csv_path.name} left as deposited")


# ------------------------------ Table S4 ---------------------------------

def edit_table_s4(csv_path: Path):
    # Original rows:
    # Unharmonized (matched 15-type PCA), 15, -0.139, 0.621
    # Ontology fix (exclude T cell),       14, -0.037, 0.899
    # Ontology + tissue restriction,        12, -0.042, 0.897
    # Full harmonization (...),             12, -0.042, 0.897
    rows = [
        ["Harmonization level", "n_types", "Spearman rho", "p-value"],
        ["Original (35-type vs 15-type PCA mismatch; Table S3)", 15, -0.386, 0.156],
        ["Level 0 (matched 15-type PCA baseline; Figure S4A)", 15, -0.139, 0.621],
        ["Level 1 (exclude T cell: tissue composition heterogeneity)", 14, -0.037, 0.899],
        ["Level 2 (Level 1 + tissue restriction)", 12, -0.042, 0.897],
        ["Level 3 (Level 2 + cell count cap; identical to Level 2)", 12, -0.042, 0.897],
    ]
    with open(csv_path, "w", newline="") as f:
        csv.writer(f).writerows(rows)
    print(f"  S4 edited: {csv_path}")


# ------------------------------ Table S5 ---------------------------------

def edit_table_s5(csv_path: Path):
    rows = list(csv.reader(open(csv_path)))
    header = rows[0]
    data = rows[1:]
    # Find plasma cell row; it should be last
    plasma_idx = next((i for i, r in enumerate(data)
                       if r and r[0].strip() == "plasma cell"), None)
    if plasma_idx is None:
        raise RuntimeError("plasma cell row not found in Table S5")
    plasma_row = data.pop(plasma_idx)
    # Insert between MSC (rank 13) and MSC of adipose (rank 11).
    # In the pre-insert table, MSC is at data index 22 (row 24 in 1-indexed
    # file) and MSC of adipose at data index 23 (row 25). After popping
    # plasma from the end, MSC is still at index 22. We want plasma
    # BETWEEN them, so insert at index 23.
    msc_idx = next((i for i, r in enumerate(data)
                    if r[0].strip() == "mesenchymal stem cell"), None)
    if msc_idx is None:
        raise RuntimeError("mesenchymal stem cell row not found")
    data.insert(msc_idx + 1, plasma_row)

    with open(csv_path, "w", newline="") as f:
        csv.writer(f).writerows([header] + data)
    print(f"  S5 edited: {csv_path}  (plasma moved from end to position after "
          f"mesenchymal stem cell, rank 12)")


# ------------------------------ Table S6 ---------------------------------

# Terminology: the per-type score is reported in the prose as the
# Procrustes residual / per-type divergence, so the column names follow that
# terminology. D1 = Sheet 2 rigidity_rank; D2 = Sheet 3 rigidity_rank_primary
# (which has no other tracked literal — renamed here so the de-rigidified name
# is reproducible from tracked code rather than a manual-only xlsx edit).
DERIG_HEADERS = {"rigidity_rank": "divergence_rank",
                 "rigidity_rank_primary": "divergence_rank_primary"}
DIVERGENCE_RANK_COL = "divergence_rank"


def _derigidify_s6_headers(wb) -> int:
    """Rename any rigidity_* column header to its divergence_* equivalent across all
    sheets. Idempotent — safe to re-run; renames in place when the column already
    exists (the current materialized state) and is a no-op on a freshly built sheet."""
    n = 0
    for ws in wb.worksheets:
        for c in range(1, ws.max_column + 1):
            h = ws.cell(1, c).value
            if h in DERIG_HEADERS:
                ws.cell(1, c).value = DERIG_HEADERS[h]
                n += 1
    return n


def edit_table_s6(xlsx_path: Path):
    # Sheet 1 has a 'rank' column per cell type. Sheet 2 is sorted by that rank with a
    # divergence_rank column for self-containment; Sheet 3 carries divergence_rank_primary.
    wb = openpyxl.load_workbook(xlsx_path)
    n_renamed = _derigidify_s6_headers(wb)

    s1 = wb["CPC1_summary"]
    rank_map = {}
    for r in range(2, s1.max_row + 1):
        ct = s1.cell(r, 1).value
        rk = s1.cell(r, 2).value
        if ct and rk is not None:
            rank_map[ct] = rk

    s2 = wb["CPC1_full_loadings"]
    header = [s2.cell(1, c).value for c in range(1, s2.max_column + 1)]
    if DIVERGENCE_RANK_COL not in header:
        # Fresh sheet (no rank column yet): add divergence_rank and sort by it.
        body = []
        for r in range(2, s2.max_row + 1):
            row = [s2.cell(r, c).value for c in range(1, s2.max_column + 1)]
            if any(v is not None for v in row):
                body.append(row)

        new_header = header + [DIVERGENCE_RANK_COL]
        new_body = [row + [rank_map.get(row[0])] for row in body]

        # Sort by (divergence_rank, species, loading_rank). Missing rank → end.
        def sort_key(row):
            rk = row[-1] if row[-1] is not None else 10**9
            species = row[1] or ""
            loading_rank = row[3] if isinstance(row[3], int) else 10**9
            return (rk, species, loading_rank)
        new_body.sort(key=sort_key)

        s2.delete_rows(1, s2.max_row)
        for c, h in enumerate(new_header, start=1):
            s2.cell(1, c).value = h
        for r_i, row in enumerate(new_body, start=2):
            for c_i, v in enumerate(row, start=1):
                s2.cell(r_i, c_i).value = v

    wb.save(xlsx_path)
    print(f"  S6 edited: {xlsx_path}  (de-rigidified {n_renamed} rigidity_* header(s); "
          f"divergence_rank ensured on Sheet 2)")


# ------------------------------ key_resources_table ---------------------

# RETIRED: edits the Key Resources Table, removed in the PLOS reformat (D5 / WP-E.3); its target (key_resources_table.md) no longer exists.
def edit_key_resources(path: Path):
    text = path.read_text()
    # Exactly one Qu scRNA-seq reference per Decision C
    old = "| Qu et al. 2022 macaque atlas | Qu et al., 2022 | GEO: GSE196791 |"
    new = "| Qu et al. 2022 macaque atlas | Qu et al., 2022 | GEO: GSE196792 |"
    if old not in text:
        raise RuntimeError("Expected GSE196791 row in key_resources_table.md not found")
    text = text.replace(old, new, 1)
    path.write_text(text)
    print(f"  key_resources_table.md: GSE196791 → GSE196792 (Qu scRNA-seq)")


# ------------------------------ Orchestrator ----------------------------

def main():
    # Edit canonical copies in supplementary_materials/
    edit_table_s1(CANON / "table_S1.xlsx")
    edit_table_s2(CANON / "table_S2.xlsx")
    edit_table_s3(CANON / "table_S3.csv")
    edit_table_s4(CANON / "table_S4.csv")
    edit_table_s5(CANON / "table_S5.csv")
    edit_table_s6(CANON / "Table_S6_CPC1_driver_genes.xlsx")

    # Staging copies (Table_S1.xlsx ... Table_S6.xlsx in figures_for_review/) are
    # materialized by scripts/build_submission_packet.py (R21 build script).
    # Run that after this script to refresh the packet.
    print("  canonical edits complete; run scripts/build_submission_packet.py to refresh packet copies")

    # key_resources_table.md (Decision C)
    edit_key_resources(DOCS / "submission/key_resources_table.md")


if __name__ == "__main__":
    main()
