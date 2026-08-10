#!/usr/bin/env python3
"""
Stage-3a Table 1 integrity + de-rigidify (reproducible synthesis pass).

S13 Table (docs/supplementary_materials/table_S13_test_inventory.xlsx) is the tracked
artifact; it has no from-scratch repro-pipeline producer (the historical builders
live in the gitignored scripts/archive/). This tracked pass applies the Stage-3a
edits in place, idempotently, so the result is reproducible from tracked code.

Changes (k = 52 unchanged; no rows added/removed):
  B  de-rigidify: "rigidity" -> "Procrustes residual" in the 7 descriptions that
     contain it (T33, T36, T37, T40, T53, T60, T61).
  C1 T59 value -0.393 -> -0.410, raw p 0.087 -> 0.073 (Stage-0 V6 recompute).
  C2 T59 Figure/Table "Fig S3D" -> "Fig S3B" (Figure S3 has only panels A, B).
  C3 T53 value +0.247 -> -0.247 (Stage-2b SAMap recompute vs Procrustes residual).
  C5 T44 Test Type "Spearman + FDR" -> "Fisher's exact + FDR" (the test is Fisher's).
  C6 footnote: state the exclusion principle for the k = 52 family.
  D22 reconcile T34/T35 to the manuscript Table 1: description -> overall 500-gene
     identity-set CellMarker enrichment (primary / expression-matched background),
     n -> "500 genes" (was the pre-d19 "top 50 genes x 6 types" / "6 types" labels),
     and add the T34/T35 caption note. Values (4.49, 3.32) and p-values unchanged.
  D28 Minor 1: T11 corrected-p (Bonferroni, col I) cell "significant" ->
     "direction-robust (100/100; resampling)", matching the body / Methods /
     Fig 3E framing of the +0.159 donor-split delta as direction-robustness, and
     the manuscript Table 1 source row. (Minor 3 / tiny-p sci-notation display was
     already correct in-repo: T34/T35/T38 raw + Bonferroni cells are 0.00E+00.)
Bonferroni columns for T59/T53 stay at 1.0 (p*52 > 1, capped) — no change.
"""
from pathlib import Path
import re
import openpyxl

XLSX = Path(__file__).resolve().parent.parent / "docs/supplementary_materials/table_S13_test_inventory.xlsx"
COL = {"id": 1, "desc": 3, "testtype": 4, "n": 5, "value": 7, "rawp": 8,
       "status": 10, "fig": 11}
SHEET_NAME = "S13 Table"        # matches the manuscript caption and the file name
SHEET_NAME_LEGACY = "Table 1"   # pre-D67 tab name, still what an unmigrated copy carries


def main():
    wb = openpyxl.load_workbook(XLSX)
    # ---- D67: the worksheet is named for the display item it is ----
    # The file, the manuscript caption (manuscript_combined.txt:330 "S13 Table."),
    # the legend and this script's own docstring all say S13 Table; the sheet tab
    # said "Table 1", which is the first thing a reader opening the workbook sees.
    # Keyed on the old name so this runs against the current tracked workbook, and
    # a fixed point once renamed.
    ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb[SHEET_NAME_LEGACY]
    ws.title = SHEET_NAME

    # ---- D57: append the three tests the paper reports but the sheet omitted ----
    # Runs before the row scan below, so everything downstream sees 64 rows.
    #
    # openpyxl.insert_rows moves cell VALUES and nothing else: merged ranges stay
    # where they were, row heights stay with their old row numbers, and new rows
    # carry no fill. Heights and the footnote merge are re-derived at the end of
    # this script from a fresh scan, so they repair themselves; the stale merge and
    # the missing fills do not, and are handled explicitly here.
    #
    # Appended rather than placed with their topical neighbours. The sheet is an
    # inventory indexed by identifier, its numbering is already non-contiguous, and
    # inserting mid-table would shift every row below for a presentational gain.
    # T67-T69 rather than the five vacant IDs, because the footnote explains those
    # five as cut analyses and reusing one would make that sentence false.
    NEW_ROWS = [
        # T67: the marker-similarity sweep contributes ONE row, at the crossover.
        # Its sixteen K values are one analysis at increasing granularity, not
        # sixteen independent tests; entering all of them would inflate k by fifteen
        # for the single claim Results section 1 makes ("not significant at fifteen").
        # Values: analysis/sensitivity_analyses/markernull_results.json,
        # ward_sweep.15.obs_null_100k = 0.9630918055960713, p_100k = 0.08586914130858692
        ("T67", "Global coherence",
         "Marker-similarity-stratified null (Ward K = 15)",
         "Label permutation (stratified)", 35, "obs/null", 0.963, 0.0859,
         None, "Confirmatory", "S5 Fig A"),
        # T68: the same file's monotonicity block, whose key name does not describe
        # what it holds -- it is the n = 35 per-type correlation, not a monotonicity
        # test over K. spearman_rho = -0.13613445378151262, spearman_p = 0.435524194238914
        ("T68", "Ranking validation",
         "Per-type residual vs marker-distinctness",
         "Spearman", 35, "ρ", -0.136, 0.4355,
         None, "Exploratory", "S5 Fig B"),
        # T69: the pre-rotation arm of the test whose post-rotation arm is T66.
        # output/twolayer_pansci_replication/pansci_layer2_summary.json,
        # layer2_pre_rotation.k5.S = 0.39642953872680664, p = 9.999000099990002e-05
        ("T69", "Two-layer",
         "Layer 2 PanSci replication (Krzanowski S, k=5, pre-rotation)",
         "Label permutation", 16, "Krzanowski S", 0.396, 0.0001,
         None, "Confirmatory", "S2 Text"),
    ]
    _have = {str(ws.cell(r, COL["id"]).value or "").strip()
             for r in range(1, ws.max_row + 1)}
    if NEW_ROWS[0][0] not in _have:
        _last_t = max(r for r in range(1, ws.max_row + 1)
                      if re.match(r"T\d+$", str(ws.cell(r, COL["id"]).value or "").strip()))
        _template = _last_t                      # T66, for its fill
        _fills = [ws.cell(_template, c)._style for c in range(1, 12)]
        _stale = [str(m) for m in ws.merged_cells.ranges]
        for _m in _stale:
            ws.unmerge_cells(_m)
        ws.insert_rows(_last_t + 1, len(NEW_ROWS))
        for _i, _row in enumerate(NEW_ROWS):
            _r = _last_t + 1 + _i
            for _c, _v in enumerate(_row, start=1):
                ws.cell(_r, _c).value = _v
                ws.cell(_r, _c)._style = _fills[_c - 1]
        # Re-merge at the shifted positions. The footnote is found by content, and
        # the second range trailed it by the same offset, so both move together.
        for _m in _stale:
            _rng = openpyxl.utils.cell.range_boundaries(_m)
            ws.merge_cells(start_row=_rng[1] + len(NEW_ROWS), start_column=_rng[0],
                           end_row=_rng[3] + len(NEW_ROWS), end_column=_rng[2])

    rowmap = {}
    footnote_row = None
    for r in range(1, ws.max_row + 1):
        cid = ws.cell(r, COL["id"]).value
        if cid is None:
            continue
        cid = str(cid).strip()
        if re.match(r"T\d+$", cid):
            rowmap[cid] = r
        elif cid.startswith("Bonferroni correction applied"):
            footnote_row = r

    # ---- B: de-rigidify descriptions ("rigidity" -> "Procrustes residual") ----
    derig = 0
    for cid, r in rowmap.items():
        d = ws.cell(r, COL["desc"]).value
        if d and "rigidity" in d.lower():
            ws.cell(r, COL["desc"]).value = re.sub(r"rigidity", "Procrustes residual", d, flags=re.I)
            derig += 1
    assert derig in (0, 7), f"expected 0 (idempotent re-run) or 7 rigidity descriptions, found {derig}"

    # ---- C1: T59 value + raw p ----
    ws.cell(rowmap["T59"], COL["value"]).value = -0.410
    ws.cell(rowmap["T59"], COL["rawp"]).value = 0.073
    # ---- C2: T59 panel ref ----
    # Superseded by the D56 entry in FIG_SET below, which normalises this to the
    # manuscript's number-first order ("S3 Fig B"). Kept as the changelog record of
    # the panel correction; the assignment itself is now made once, in FIG_SET.
    # ---- C3: T53 SAMap sign ----
    ws.cell(rowmap["T53"], COL["value"]).value = -0.247
    # ---- C5: T44 test method ----
    tt = ws.cell(rowmap["T44"], COL["testtype"]).value
    if tt and "spearman" in str(tt).lower():
        ws.cell(rowmap["T44"], COL["testtype"]).value = "Fisher's exact + FDR"

    # ---- D61: mechanistic nulls 1-6 are pre-specified, so Confirmatory ----
    # S1_Text.txt:60 states the design status of the ten mechanistic nulls: "nulls
    # 1-6 pre-specified as confound diagnostics; 7-9 designed sequentially with
    # rho < 0.35 closure thresholds locked before invocation; 10 exploratory." The
    # footnote's own definition is Confirmatory = pre-specified or direct
    # replication, so the six pre-specified nulls belong there; all six said
    # Exploratory.
    #
    # T48-T50 (nulls 7-9) stay Exploratory deliberately. Locking a closure
    # threshold before invocation pre-commits the decision rule, not the
    # hypothesis, and a hypothesis designed after the previous result was seen is
    # not pre-specified whatever the discipline of its thresholds. They are
    # pre-registered in their analysis and post-hoc in their existence. T51
    # (null 10) is called exploratory by S1 Text itself.
    #
    # k does not move: family membership is set by the em-dash in the Bonferroni
    # column, and none of T42-T47 is among the nine excluded IDs.
    for _tid in ("T42", "T43", "T44", "T45", "T46", "T47"):
        ws.cell(rowmap[_tid], COL["status"]).value = "Confirmatory"

    # ---- C6: exclusion-principle footnote wording ----
    if footnote_row:
        f = ws.cell(footnote_row, COL["id"]).value
        f = f.replace(
            "Bonferroni correction applied over k = 52 inferential tests "
            "(excluding T03, T04, T09, T10, T18, T62, T63, T64, T65 — descriptive/exploratory, aggregate, or diagnostic).",
            "Bonferroni correction applied over the k = 52 principal inferential tests "
            "(k = 52 = 61 − 9). The 9 excluded IDs (T03, T04, T09, T10, T18, T62, T63, T64, T65) are descriptive, "
            "aggregate, or diagnostic quantities that make no inferential significance claim and are therefore not "
            "part of the multiple-testing family.")
        ws.cell(footnote_row, COL["id"]).value = f

    # ---- D22: reconcile T34/T35 to the manuscript Table 1 (idempotent) ----
    # The manuscript (d19) describes the inventoried CellMarker enrichments as the
    # overall 500-gene identity set (n = 500 genes); the spreadsheet retained the
    # pre-d19 "top 50 genes x 6 types" / "6 types" labels. Reconcile description + n
    # and add the T34/T35 caption note. Values (4.49, 3.32) + p-values unchanged.
    ws.cell(rowmap["T34"], COL["desc"]).value = (
        "Overall 500-gene identity-set CellMarker enrichment (primary 16,959-gene background)")
    ws.cell(rowmap["T35"], COL["desc"]).value = (
        "Overall 500-gene identity-set CellMarker enrichment (expression-matched background)")
    for _tid in ("T34", "T35"):
        ws.cell(rowmap[_tid], COL["n"]).value = "500 genes"
    if footnote_row:
        _fv = ws.cell(footnote_row, COL["id"]).value
        # D61: the sentence is keyed on the SUPERSEDED text, not on the "T34/T35:"
        # sentinel. The tracked footnote already satisfies that sentinel, so editing
        # the literal under the old guard would have changed nothing while leaving
        # every gate green -- a silent no-op reporting success.
        _t3435 = (
            "T34/T35: the inventoried fold-enrichments are for the overall 500-gene "
            "identity set; the per-type top-50 centroid-deviation test on the 6 validation "
            "types is a pass-rate rather than a single inventory p-value, and is not "
            "reported in this paper. ")
        # D61: the superseded sentence said the pass-rate "is reported as", gave two
        # numbers for it, and cited Figure S6 -- a display item cut from the paper
        # (reproduce/validate.py annotates its three CellMarker checks "old Fig S6,
        # cut"). The result appears in none of the three submitted texts, so S13 was
        # the only place in the submission carrying those numbers.
        _t3435_d22 = (
            "T34/T35: the inventoried fold-enrichments are for the overall 500-gene "
            "identity set; the per-type top-50 centroid-deviation test on the 6 validation "
            "types is reported as a pass-rate (5 of 6 under CellMarker, 6 of 6 under the "
            "held-out HPA reference; Figure S6, Methods), not as a single inventory p-value. ")
        if _t3435_d22 in _fv:
            _fv = _fv.replace(_t3435_d22, _t3435, 1)
            ws.cell(footnote_row, COL["id"]).value = _fv
        elif "T34/T35:" not in _fv:
            _anchor = "T29, T54, and the T55-T57 treeness tests"
            _fv = _fv.replace(_anchor, _t3435 + _anchor, 1)
            ws.cell(footnote_row, COL["id"]).value = _fv

    # ---- D55: state the counting rule on the sheet itself (idempotent) ----
    # k = 52 was derivable only from the footnote arithmetic. The Bonferroni column
    # already marks the nine exclusions with an em-dash, exactly and only those nine,
    # so the rule needs stating rather than encoding. The second sentence names the
    # one row that does not look like it follows the rule: T11 is inside the family
    # but reports a resampling CI, so it carries no corrected p. T11's cell is
    # correct as it stands and is not edited.
    if footnote_row:
        _fv = ws.cell(footnote_row, COL["id"]).value
        if "carrying an em-dash in the Bonferroni" not in _fv:
            _rule = (
                " The nine excluded IDs are the rows carrying an em-dash in the "
                "Bonferroni p column. T11 is inside the family of 52 but reports a "
                "resampling confidence interval rather than a p-value, so it carries "
                "no corrected p; its direction robustness is given in that column "
                "instead.")
            _anchor = "the inventory totals 61 tests."
            _fv = _fv.replace(_anchor, _anchor + _rule, 1)
            ws.cell(footnote_row, COL["id"]).value = _fv

    # ---- D28: Minor 1 — T11 corrected-p (Bonferroni, col I) label (idempotent) ----
    # The donor-split delta (+0.159) is framed in the body / Methods / Fig 3E caption
    # / Limitations as direction-robustness (100/100 resampling splits), not a
    # calibrated effect. Align the table's corrected-p cell with that framing and with
    # the manuscript Table 1 source row. Column I (9) = "Bonferroni p (k=52)".
    ws.cell(rowmap["T11"], 9).value = "direction-robust (100/100; resampling)"

    # ---- D55: figure-column repair against the five-figure paper (idempotent) ----
    # This map replaces the D3 seven-figure map. That map named Fig 6B, Fig 7A and
    # Fig 7B, none of which exists in the re-assembled paper, and it pointed several
    # rows at panels that do exist but now show something else -- Fig 4C is the
    # recovery ceiling, not Layer 2. Thirty-eight of the sixty-one rows move.
    #
    # EM_DASH means "no display item", which is the truthful entry for a test whose
    # result appears in no submitted text. Eight rows take it. Six of those eight are
    # Confirmatory and stay inside k = 52: a multiple-comparison family counts tests
    # performed, not tests reported.
    EM_DASH = "—"
    FIG_SET = {
        # text-only results: bootstrap CV and LOOCV (S1 §11), donor split (S1 §9),
        # macaque (S1 §7), the ten mechanistic nulls (S1 §8)
        "T03": "S1 Text", "T04": "S1 Text",
        "T09": "S1 Text", "T10": "S1 Text", "T11": "S1 Text",
        "T12": "S1 Text", "T14": "S1 Text", "T15": "S1 Text",
        "T42": "S1 Text", "T43": "S1 Text", "T44": "S1 Text", "T45": "S1 Text",
        "T46": "S1 Text", "T47": "S1 Text", "T48": "S1 Text", "T49": "S1 Text",
        "T50": "S1 Text", "T51": "S1 Text",
        # missing callouts the current display items do carry
        "T13": "Fig 1D",      # the human-mouse-lemur null IS Fig 1D
        "T66": "S2 Text",     # S2 Text reports S = 0.402 against null 0.360
        # panels that moved under the seven-to-five renumbering
        "T19": "Fig 3", "T20": "Fig 3", "T21": "Fig 3",   # Fig 3 is one panel
        "T30": "Fig 2B",      # Layer 2 pre/post rotation
        # no display item: the result appears in no submitted text
        "T31": EM_DASH, "T34": EM_DASH, "T35": EM_DASH, "T36": EM_DASH,
        "T37": EM_DASH, "T38": EM_DASH, "T39": EM_DASH, "T53": EM_DASH,

        # ---- D56: supplementary references, number first ----
        # The submitted manuscript and both supporting texts write supplementary
        # display items number first -- S1 Fig, S1 Table -- in 52 places between
        # them, and never write Fig S1 or Table S1 anywhere. Table 1 used the
        # inverted form in these eighteen cells, so a reader following a callout
        # had to translate it. Main-figure references already matched (the
        # manuscript writes Fig 4C) and are untouched above.
        #
        # Panel letters are kept rather than dropped. The submitted texts use them
        # nowhere in prose, but a table column exists to carry that precision, and
        # S1 Fig A follows the number-first order the texts do use.
        "T05": "S1 Fig A", "T06": "S1 Fig B",
        "T63": "S1 Fig F", "T64": "S1 Fig D", "T65": "S1 Fig E",
        "T18": "S2 Fig E",
        "T22": "S2 Fig F", "T23": "S2 Fig F", "T24": "S2 Fig F",
        "T60": "S2 Fig C", "T61": "S2 Fig C", "T62": "S2 Fig C",
        "T59": "S3 Fig B",
        "T16": "S4 Fig", "T17": "S4 Fig",     # matched-scale; no S7 Fig exists
        "T40": "S1 Table",
        "T58": "S3 Table",    # S3 Table carries rho -0.526, p 0.044, n 15
        "T27": "S4 Table",    # S4 Table carries rho -0.139, p 0.621, n 15
    }
    for tid, val in FIG_SET.items():
        ws.cell(rowmap[tid], COL["fig"]).value = val

    # ---- D55: T64 ranking recovery ceiling, superseded by commit 4615491 ----
    # 0.42 contradicted S1 Fig D's own caption, which reads "peaks at rho ~ 0.45 in
    # the calibrated-signal regime". The description names the regime because 0.42
    # remains correct for the deposited grid at signal 3.0 and the two must be
    # distinguishable. The Fig S1D callout is correct and does not change.
    ws.cell(rowmap["T64"], COL["value"]).value = 0.45
    ws.cell(rowmap["T64"], COL["desc"]).value = (
        "Ranking recovery ceiling (calibrated signal)")

    # ---- D55: T52 n, resolved from the producer ----
    # output/t3g/primary_correlation_results.json carries both drug-target tests.
    # T51 is primary_correlation (conservation ratio), n = 34, defined only for the
    # 34 types that have drug targets. T52 is density_correlation (drug-target
    # density), n = 35, defined for all types because a density may be zero. The
    # sheet had propagated T51 n to T52.
    ws.cell(rowmap["T52"], COL["n"]).value = 35

    # ---- D57: k = 52 -> 55, header and the whole corrected-p column ----
    # Adding three in-family rows recomputes EVERY corrected p, not just the new
    # ones. No row changes verdict: no in-family raw p falls in the band between
    # 0.05/55 and 0.05/52, and the nearest values are two orders of magnitude away
    # on either side.
    #
    # Three classes, not one. A bare float(raw) * k raises on the four rows whose
    # raw p is prose.
    #   numeric      multiply, cap at 1, following the sheet's existing convention
    #   bounded <x   multiply the bound, keep the prefix, as T01 already does
    #   prose        leave untouched; the correction factor does not apply
    K_NEW = 55
    BONF_COL = 9
    EM_DASH_BONF = "—"      # marks the nine excluded rows; must stay at nine
    _hdr = ws.cell(1, BONF_COL).value
    ws.cell(1, BONF_COL).value = _hdr.replace("k=52", "k=%d" % K_NEW)

    # The sheet's existing convention is the EXACT product, not a rounded one:
    # 0.0043 x 52 is stored as 0.2236 and 2.1e-13 x 52 as 1.092e-11. Multiplying in
    # binary floating point would store 0.013 x 52 as 0.6759999999999999, so the
    # arithmetic is done in Decimal from the raw cell's own digits, which reproduces
    # the convention exactly and leaves no noise to clean up afterwards.
    from decimal import Decimal

    def _product(raw_text):
        return Decimal(raw_text) * K_NEW

    def _fmt(d):
        """Capped at 1, else the exact product with the sheet's exponent style."""
        if d > 1:
            return 1
        s = format(d.normalize(), "f") if Decimal("1e-4") <= abs(d) else None
        if s is not None:
            return float(s)
        # small values keep scientific notation, exponent written without a
        # leading zero, as the sheet already writes 5.2e-5 and 1.092e-11
        return float(d)

    _excluded = {"T03", "T04", "T09", "T10", "T18", "T62", "T63", "T64", "T65"}
    _counts = {"numeric": 0, "bounded": 0, "prose": 0, "excluded": 0}
    for _tid, _r in rowmap.items():
        _cell = ws.cell(_r, BONF_COL)
        _cur = str(_cell.value).strip()
        if _cur == EM_DASH_BONF:
            _counts["excluded"] += 1
            continue
        _raw = str(ws.cell(_r, COL["rawp"]).value).strip()
        _m = re.fullmatch(r"<\s*([\d.eE+-]+)", _raw)
        if _m:
            _p = _product(_m.group(1))
            # "< 5.2e-5", not "< 5.2e-05": strip the exponent's leading zero so the
            # new bounds read as the existing ones do.
            _cell.value = "< %s" % ("%g" % float(_p)).replace("e-0", "e-")
            _counts["bounded"] += 1
            continue
        try:
            _p = _product(_raw)
        except Exception:
            _counts["prose"] += 1          # T11, T44, T47, T48: leave as-is
            continue
        _cell.value = _fmt(_p)
        _counts["numeric"] += 1
    print("  Bonferroni recompute at k=%d: %s" % (K_NEW, _counts))

    # ---- D57: the footnote, restated for 64 tests and a family of 55 ----
    # The D55 sentence said "the family of 52" and "totals 61 tests". Both move.
    #
    # It also named T11 as the row inside the family carrying no corrected p, and
    # that remains exactly right: T11's corrected-p cell is the only prose one in
    # the sheet. Three further rows (T44, T47, T48) have a non-numeric RAW p but do
    # carry a corrected p of 1, so they are not exceptions to the counting rule.
    # The sentence gains a clause distinguishing the two situations, because the
    # difference is invisible from the corrected-p column alone.
    if footnote_row:
        _fv = ws.cell(footnote_row, COL["id"]).value
        _fv = _fv.replace("k = 52 = 61 − 9", "k = 55 = 64 − 9")
        _fv = _fv.replace("k = 52 principal inferential tests",
                          "k = 55 principal inferential tests")
        _fv = _fv.replace("the inventory totals 61 tests.",
                          "the inventory totals 64 tests.")
        _fv = _fv.replace("T11 is inside the family of 52 but",
                          "T11 is inside the family of 55 but")
        _sub = (" T44, T47 and T48 also report a raw p as a range or a bound over "
                "several sub-tests rather than as a single value, but each is "
                "corrected and capped like any other in-family row.")
        # D68: this guard was keyed on "T44, T47 and T48 also report", and D68's F2
        # rewrites that clause to drop the "also". Left as it was, the sentinel would
        # stop matching after D68 ran, this block would fire on the NEXT run, and it
        # would re-append the superseded sentence after the "given in that column
        # instead." anchor that F2 keeps -- so the script would stop being a fixed
        # point and the footnote would grow a duplicate. Keyed now on the substring
        # both wordings share, which still fires on a footnote missing the clause.
        if "T44, T47 and T48" not in _fv:
            _anchor = "given in that column instead."
            _fv = _fv.replace(_anchor, _anchor + _sub, 1)
        ws.cell(footnote_row, COL["id"]).value = _fv

    # ---- D68: state the membership rule, and name its boundary cases ----
    # The footnote said the nine excluded IDs "are descriptive, aggregate, or
    # diagnostic quantities that make no inferential significance claim". That is
    # true of four of them. T03, T04, T09, T10 and T18 are excluded for a different
    # and more general reason: they report a pass rate over repeated sub-tests, not
    # one p-value for one hypothesis. Verified against the sheet before writing --
    # their Raw p cells read "100/100 sig", "35/35 < 1.0", "all p < 0.01 (100/100)"
    # twice, and "20/24 sig". Six of the nine (T03, T04, T09, T10, T18, T62) are
    # Confirmatory, so the old sentence also implied a status the sheet contradicts.
    # Stating the rule makes the nine derivable rather than merely listed.
    #
    # SENTINEL. Keyed on "Membership rule:", which exists only in the text this
    # block writes. Keying it on anything already present in the tracked footnote
    # would make the block a silent no-op: it would report success, leave the
    # workbook bytes untouched, and hold TABLE_1_LOCK_MD5 and all four gates green
    # while the edit had not landed. That is not hypothetical -- see the D61 note at
    # the T34/T35 block above, where exactly that nearly shipped. The replacements
    # below also raise rather than skip if a target is missing, so a footnote that
    # has drifted fails loudly instead of half-applying.
    if footnote_row:
        _fv = ws.cell(footnote_row, COL["id"]).value
        if "Membership rule:" not in _fv:
            _f1_old = (
                "The 9 excluded IDs (T03, T04, T09, T10, T18, T62, T63, T64, T65) are "
                "descriptive, aggregate, or diagnostic quantities that make no "
                "inferential significance claim and are therefore not part of the "
                "multiple-testing family.")
            _f1_new = (
                "Membership rule: a test joins the family when it yields a single "
                "p-value for a single hypothesis. The 9 excluded IDs do not. T03, T04, "
                "T09, T10 and T18 report a pass rate over repeated sub-tests (for "
                "example 100 of 100 subsamples significant) rather than one p-value for "
                "one hypothesis; their Raw p column gives that pass rate. T62, T63, T64 "
                "and T65 are descriptive or simulation-diagnostic quantities that make "
                "no inferential claim. Exclusion is a statement about the form of the "
                "result, not about whether the row is informative: several excluded rows "
                "are Confirmatory in the sense of the Status column.")
            _f2_old = (
                "T11 is inside the family of 55 but reports a resampling confidence "
                "interval rather than a p-value, so it carries no corrected p; its "
                "direction robustness is given in that column instead. T44, T47 and T48 "
                "also report a raw p as a range or a bound over several sub-tests rather "
                "than as a single value, but each is corrected and capped like any other "
                "in-family row.")
            _f2_new = (
                "Four rows are exceptions to the membership rule in the other direction, "
                "retained in the family by judgement and named here rather than left to "
                "inference. T11 reports a resampling confidence interval rather than a "
                "p-value, so it carries no corrected p; its direction robustness is given "
                "in that column instead. T44, T47 and T48 report a raw p as a range or a "
                "bound over several sub-tests rather than as a single value, but each is "
                "corrected and capped like any other in-family row.")
            for _old, _new in ((_f1_old, _f1_new), (_f2_old, _f2_new)):
                if _old not in _fv:
                    raise SystemExit(
                        "D68: footnote target absent, refusing to half-apply: %r" % _old[:70])
                _fv = _fv.replace(_old, _new, 1)
            ws.cell(footnote_row, COL["id"]).value = _fv

    # ---- D61: the last inverted supplementary reference in the workbook ----
    # D56 normalised the 18 references in column K to the number-first order the
    # submitted texts use; LEGACY_SUPP_RE is anchored to a whole cell, so it could
    # not reach a reference embedded in footnote prose. This one is the remainder.
    # No block of this script has ever authored the sentence -- it is inherited
    # workbook content -- so it needs its own edit rather than a corrected literal.
    if footnote_row:
        _fv = ws.cell(footnote_row, COL["id"]).value
        _fv = _fv.replace("documented in Table S4 rather than",
                          "documented in S4 Table rather than")
        ws.cell(footnote_row, COL["id"]).value = _fv

    # ---- D61: derive the row fill from Status instead of writing it alongside ----
    # The footnote states the rule -- "Green rows = confirmatory; yellow rows =
    # exploratory or diagnostic" -- and the sheet obeyed it for 63 of 64 rows. T68
    # was the exception, green on an Exploratory row: when D57 appended T67-T69 it
    # copied the template row's style, and green happened to be right for T67 and
    # T69 and wrong for T68. A hand-maintained convention that agrees 63 times out
    # of 64 is worth deriving rather than repairing, so it is derived here and no
    # row's fill can contradict its Status again.
    #
    # The assert is the point of the block as much as the fill is: an unrecognised
    # Status would otherwise fall through to yellow silently.
    from openpyxl.styles import PatternFill
    _GREEN = PatternFill("solid", fgColor="00E2EFDA", bgColor="00000000")
    _YELLOW = PatternFill("solid", fgColor="00FFF2CC", bgColor="00000000")
    for _tid, _r in rowmap.items():
        _st = str(ws.cell(_r, COL["status"]).value or "")
        assert _st.startswith(("Confirmatory", "Exploratory", "Diagnostic")), \
            f"{_tid}: unrecognised Status {_st!r}; fill cannot be derived"
        _fill = _GREEN if _st.startswith("Confirmatory") else _YELLOW
        for _c in range(1, 12):
            ws.cell(_r, _c).fill = _fill

    # ---- Stage-5.5 C: presentation formatting (reproducible cell sizing) ----
    # Column widths so the Description / Status / Category cells are not clipped.
    from openpyxl.styles import Alignment
    widths = {"A": 6, "B": 18, "C": 72, "D": 24, "E": 7, "F": 11,
              "G": 11, "H": 13, "I": 14, "J": 27, "K": 14}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    # Wrap text + top-align the prose columns; size each data row to its lines.
    wrap_cols = {3: 72, 10: 27, 2: 18, 4: 24}   # Description, Status, Category, Test Type
    for r in rowmap.values():
        n_lines = 1
        for ci, w in wrap_cols.items():
            ws.cell(r, ci).alignment = Alignment(wrap_text=True, vertical="top")
            txt = str(ws.cell(r, ci).value or "")
            n_lines = max(n_lines, -(-len(txt) // max(1, w - 2)))
        ws.row_dimensions[r].height = 15 * n_lines + 4
    # Footnote: span the full table width, wrap, and a tall row so it shows fully.
    if footnote_row:
        ws.merge_cells(start_row=footnote_row, start_column=1,
                       end_row=footnote_row, end_column=11)
        fc = ws.cell(footnote_row, 1)
        fc.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
        ws.row_dimensions[footnote_row].height = 150

    # ---- E.2 rev: deterministic save (eliminate wall-clock timestamps) ----
    from datetime import datetime as _dt
    import zipfile as _zf
    import os as _os
    _FIXED = _dt(2026, 1, 1, 0, 0, 0)
    _ISO = "2026-01-01T00:00:00Z"
    wb.properties.created = _FIXED
    wb.properties.modified = _FIXED

    wb.save(XLSX)

    # Rewrite the .xlsx zip: fixed entry date_time for every member + fixed
    # docProps/core.xml dcterms timestamps (same content/names/order/compression).
    with _zf.ZipFile(XLSX, "r") as _zin:
        _members = [(i, _zin.read(i.filename)) for i in _zin.infolist()]
    _tmp = str(XLSX) + ".tmp"
    with _zf.ZipFile(_tmp, "w") as _zout:
        for _info, _data in _members:
            if _info.filename == "docProps/core.xml":
                _s = _data.decode("utf-8")
                _s = re.sub(r"(<dcterms:created[^>]*>)[^<]*(</dcterms:created>)",
                            r"\g<1>" + _ISO + r"\g<2>", _s)
                _s = re.sub(r"(<dcterms:modified[^>]*>)[^<]*(</dcterms:modified>)",
                            r"\g<1>" + _ISO + r"\g<2>", _s)
                _data = _s.encode("utf-8")
            _ni = _zf.ZipInfo(_info.filename, date_time=(2026, 1, 1, 0, 0, 0))
            _ni.compress_type = _info.compress_type
            _ni.external_attr = _info.external_attr
            _ni.create_system = _info.create_system
            _zout.writestr(_ni, _data)
    _os.replace(_tmp, XLSX)

    # ---- verify ----
    wb2 = openpyxl.load_workbook(XLSX)
    w = wb2[SHEET_NAME]
    rig = sum(1 for r in range(1, w.max_row + 1)
              if w.cell(r, COL["desc"]).value and "rigidity" in str(w.cell(r, COL["desc"]).value).lower())
    print(f"  de-rigidified {derig} descriptions; remaining 'rigidity' in descriptions: {rig}")
    print(f"  T59 value={w.cell(rowmap['T59'],COL['value']).value} rawp={w.cell(rowmap['T59'],COL['rawp']).value} fig={w.cell(rowmap['T59'],COL['fig']).value}")
    print(f"  T53 value={w.cell(rowmap['T53'],COL['value']).value}")
    print(f"  T44 testtype={w.cell(rowmap['T44'],COL['testtype']).value}")
    for _tid in ("T34", "T35"):
        print(f"  {_tid} desc={w.cell(rowmap[_tid],COL['desc']).value!r} n={w.cell(rowmap[_tid],COL['n']).value!r} value={w.cell(rowmap[_tid],COL['value']).value}")
    print(f"  T11 corrected-p (I{rowmap['T11']})={w.cell(rowmap['T11'],9).value!r}")
    print(f"  T34/T35 note present: {'T34/T35:' in str(w.cell([r for r in range(1,w.max_row+1) if str(w.cell(r,1).value or '').startswith('Bonferroni correction applied')][0],1).value)}")
    print(f"  saved {XLSX}")


if __name__ == "__main__":
    main()
